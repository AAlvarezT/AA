#3er-interfaz-B3
import tkinter as tk
from tkinter import ttk
import openpyxl
from datetime import datetime

# Ruta del archivo Excel
archivo_excel = r'C:\Users\ADMIN\Desktop\PROYECTO_BASE_3\Plantilla_B3.xlsx'

# Función para cargar los clientes que hicieron pedidos en la fecha actual
def cargar_clientes_por_fecha():
    # Obtener la fecha actual en formato "YYYY-MM-DD"
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    fecha_entry.config(state="normal")
    fecha_entry.delete(0, tk.END)
    fecha_entry.insert(0, fecha_actual)
    fecha_entry.config(state="readonly")

    # Cargar el archivo Excel y seleccionar la hoja "Pedidos"
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        hoja_pedidos = libro['Pedidos']
        hoja_clientes = libro['Clientes']

        # Obtener los nombres de los clientes que tienen pedidos en la fecha actual
        clientes_en_fecha = set()
        for fila in range(2, hoja_pedidos.max_row + 1):
            fecha_pedido = hoja_pedidos.cell(row=fila, column=3).value  # Columna 3 es Fecha
            if fecha_pedido:
                fecha_pedido = fecha_pedido.strftime("%Y-%m-%d")
            if fecha_pedido == fecha_actual:
                id_cliente = hoja_pedidos.cell(row=fila, column=2).value  # Columna 2 es ID_Cliente
                if id_cliente:
                    # Obtener el nombre del cliente según el ID_Cliente
                    nombre_cliente = hoja_clientes.cell(row=id_cliente + 1, column=2).value  # Columna 2 es Nombre
                    if nombre_cliente:
                        clientes_en_fecha.add(nombre_cliente)

        # Ordenar alfabéticamente y actualizar la lista desplegable con los nombres de los clientes
        cliente_opciones['values'] = sorted(clientes_en_fecha)
        if clientes_en_fecha:
            cliente_opciones.current(0)  # Seleccionar el primer cliente en la lista
        else:
            cliente_opciones.set('')  # No hay clientes para esa fecha

        libro.close()
    except Exception as e:
        resultado_label.config(text=f"Error al cargar datos: {e}")

# Función para guardar los datos seleccionados
def guardar_datos():
    estado = estado_var.get()
    forma_pago = forma_pago_var.get()
    nombre_cliente = cliente_opciones.get()
    
    # Cargar el archivo Excel y seleccionar la hoja "Pedidos"
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        hoja_pedidos = libro['Pedidos']
        hoja_clientes = libro['Clientes']

        # Buscar el pedido por nombre de cliente y fecha, luego actualizar los campos
        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        encontrado = False
        for fila in range(2, hoja_pedidos.max_row + 1):
            fecha_pedido = hoja_pedidos.cell(row=fila, column=3).value  # Columna 3 es Fecha
            if fecha_pedido:
                fecha_pedido = fecha_pedido.strftime("%Y-%m-%d")
            id_cliente = hoja_pedidos.cell(row=fila, column=2).value  # Columna 2 es ID_Cliente
            nombre_cliente_excel = hoja_clientes.cell(row=id_cliente + 1, column=2).value if id_cliente else None

            if fecha_pedido == fecha_actual and nombre_cliente_excel == nombre_cliente:
                hoja_pedidos.cell(row=fila, column=4, value=forma_pago)  # Columna 4 es ID_FormaPago
                hoja_pedidos.cell(row=fila, column=5, value=estado)      # Columna 5 es ID_Estado
                encontrado = True
                break

        # Guardar los cambios en el archivo Excel
        if encontrado:
            libro.save(archivo_excel)
            resultado_label.config(text="Datos guardados correctamente")
        else:
            resultado_label.config(text="No se encontró el pedido para actualizar")

        libro.close()
    except Exception as e:
        resultado_label.config(text=f"Error al guardar datos: {e}")

# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Actualizar Pedido")
ventana.geometry("350x400")

# Variable para almacenar el estado del pedido y la forma de pago
estado_var = tk.StringVar()
forma_pago_var = tk.StringVar()

# Mostrar la fecha actual automáticamente en la entrada
fecha_label = tk.Label(ventana, text="Fecha (YYYY-MM-DD):")
fecha_label.pack(pady=5)
fecha_entry = tk.Entry(ventana, state="readonly")
fecha_entry.pack(pady=5)

# Botón para cargar los clientes por la fecha actual
cargar_clientes_boton = tk.Button(ventana, text="Cargar Clientes", command=cargar_clientes_por_fecha)
cargar_clientes_boton.pack(pady=5)

# Lista desplegable para los clientes
cliente_label = tk.Label(ventana, text="Cliente:")
cliente_label.pack(pady=5)
cliente_opciones = ttk.Combobox(ventana)
cliente_opciones.pack(pady=5)

# Etiqueta y opción de selección para el estado del pedido
estado_label = tk.Label(ventana, text="Estado del Pedido:")
estado_label.pack(pady=5)
estado_opciones = ttk.Combobox(ventana, textvariable=estado_var)
estado_opciones['values'] = ("Completado", "Pendiente", "Cancelado")
estado_opciones.current(0)  # Seleccionar la primera opción por defecto
estado_opciones.pack(pady=5)

# Etiqueta y opción de selección para la forma de pago
forma_pago_label = tk.Label(ventana, text="Forma de Pago:")
forma_pago_label.pack(pady=5)
forma_pago_opciones = ttk.Combobox(ventana, textvariable=forma_pago_var)
forma_pago_opciones['values'] = ("Efectivo", "Yape")
forma_pago_opciones.current(0)  # Seleccionar la primera opción por defecto
forma_pago_opciones.pack(pady=5)

# Botón para guardar los datos
guardar_boton = tk.Button(ventana, text="Guardar", command=guardar_datos)
guardar_boton.pack(pady=10)

# Etiqueta para mostrar el resultado
resultado_label = tk.Label(ventana, text="")
resultado_label.pack(pady=10)

# Cargar los clientes para la fecha actual al iniciar
cargar_clientes_por_fecha()

# Iniciar el bucle de la aplicación
ventana.mainloop()